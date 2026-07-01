"""
GitHub AI 密钥扫描工具
用途：扫描 GitHub 仓库中暴露的大模型 API 密钥，输出 Excel 报告
作者：GGOB for ZhouXuan
"""

import requests
import re
import time
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    exit(1)

# ============================================================
# 配置区
# ============================================================

# GitHub Personal Access Token (用于提高 API 速率限制)
# 不填也能用，但每分钟只能搜 10 次；填了可以到 30 次
GITHUB_TOKEN = ""

# 扫描目标：搜索词 → (供应商名, 密钥正则)
# 正则用于二次验证，过滤掉占位符/示例
SEARCH_PATTERNS = {
    # --- 国际大模型 ---
    '"OPENAI_API_KEY=sk-"': (
        "OpenAI",
        r"sk-[A-Za-z0-9]{20,}",
        "GPT-4/GPT-4o/o1/o3"
    ),
    '"ANTHROPIC_API_KEY=sk-ant-"': (
        "Anthropic",
        r"sk-ant-[A-Za-z0-9\-]{20,}",
        "Claude 3.5/4"
    ),
    '"GOOGLE_API_KEY=AIza"': (
        "Google AI",
        r"AIza[A-Za-z0-9_\-]{30,}",
        "Gemini Pro/Ultra"
    ),
    '"COHERE_API_KEY="': (
        "Cohere",
        r"[a-zA-Z0-9]{40}",
        "Command R+"
    ),
    '"HUGGING_FACE_HUB_TOKEN=hf_"': (
        "HuggingFace",
        r"hf_[A-Za-z0-9]{20,}",
        "HF Inference"
    ),
    '"MISTRAL_API_KEY="': (
        "Mistral AI",
        r"[a-zA-Z0-9]{32}",
        "Mistral Large/Medium"
    ),
    '"GROQ_API_KEY=gsk_"': (
        "Groq",
        r"gsk_[A-Za-z0-9]{20,}",
        "Groq LPU"
    ),
    '"REPLICATE_API_TOKEN=r8_"': (
        "Replicate",
        r"r8_[A-Za-z0-9]{20,}",
        "Replicate"
    ),
    '"TOGETHER_API_KEY="': (
        "Together AI",
        r"[a-f0-9]{64}",
        "Together Inference"
    ),
    '"FIREWORKS_API_KEY=fw_"': (
        "Fireworks AI",
        r"fw_[A-Za-z0-9]{20,}",
        "Fireworks"
    ),
    '"PERPLEXITY_API_KEY=pplx-"': (
        "Perplexity",
        r"pplx-[a-f0-9]{40,}",
        "Perplexity AI"
    ),

    # --- 国产大模型 ---
    '"DEEPSEEK_API_KEY=sk-"': (
        "DeepSeek",
        r"sk-[a-f0-9]{20,}",
        "DeepSeek V2/V3/R1"
    ),
    '"MOONSHOT_API_KEY=sk-"': (
        "Moonshot (Kimi)",
        r"sk-[a-zA-Z0-9]{20,}",
        "Kimi/Moonshot"
    ),
    '"QWEN_API_KEY=sk-"': (
        "通义千问 (Qwen)",
        r"sk-[a-f0-9]{20,}",
        "Qwen-Max/Turbo"
    ),
    '"DASHSCOPE_API_KEY=sk-"': (
        "阿里云 DashScope",
        r"sk-[a-f0-9]{20,}",
        "DashScope"
    ),
    '"ZHIPUAI_API_KEY="': (
        "智谱 AI (GLM)",
        r"[a-f0-9]{32}\.[a-zA-Z0-9]{20,}",
        "GLM-4"
    ),
    '"BAICHUAN_API_KEY=sk-"': (
        "百川智能",
        r"sk-[a-f0-9]{20,}",
        "Baichuan"
    ),
    '"MINIMAX_API_KEY="': (
        "MiniMax",
        r"eyJhbGci[A-Za-z0-9\-_\.]+",
        "MiniMax abab"
    ),
    '"SPARK_API_KEY="': (
        "讯飞星火",
        r"[a-f0-9]{32}",
        "Spark"
    ),
    '"SILICONFLOW_API_KEY=sk-"': (
        "SiliconFlow",
        r"sk-[a-zA-Z0-9]{20,}",
        "SiliconFlow"
    ),
    '"YI_API_KEY="': (
        "零一万物 (01.AI)",
        r"[a-f0-9]{32}",
        "Yi-Large"
    ),
    '"STEPFUN_API_KEY=sk-"': (
        "阶跃星辰 (Step)",
        r"sk-[a-zA-Z0-9]{20,}",
        "Step"
    ),
    '"SENSETIME_API_KEY="': (
        "商汤 SenseNova",
        r"[a-f0-9]{32}",
        "SenseNova"
    ),
    '"HUNYUAN_SECRET_KEY="': (
        "腾讯混元",
        r"[a-f0-9]{32}",
        "Hunyuan"
    ),
}

# 占位符关键词（包含这些的跳过）
PLACEHOLDER_KEYWORDS = [
    "your_key_here", "your_api_key", "xxxxxx", "placeholder",
    "example", "test", "dummy", "fake", "sample", "insert",
    "replace_me", "changeme", "sk-xxx", "sk-ant-xxx",
    "your_openai", "your_anthropic", "TODO", "FIXME",
    "PASTE_YOUR", "ENTER_YOUR", "PUT_YOUR",
]

# ============================================================
# 核心逻辑
# ============================================================

def github_headers():
    h = {"Accept": "application/vnd.github.v3.text-match+json"}
    if GITHUB_TOKEN:
        h["Authorization"] = f"token {GITHUB_TOKEN}"
    return h


def search_github(query, page=1, per_page=30):
    """调用 GitHub Code Search API"""
    url = "https://api.github.com/search/code"
    params = {
        "q": query,
        "page": page,
        "per_page": per_page,
    }
    try:
        resp = requests.get(url, headers=github_headers(), params=params, timeout=15)
        if resp.status_code == 403:
            reset = int(resp.headers.get("X-RateLimit-Reset", 0))
            wait = max(reset - int(time.time()), 10)
            print(f"  ⏳ 限速，等待 {wait}s...")
            time.sleep(wait)
            return search_github(query, page, per_page)
        if resp.status_code == 422:
            print(f"  ⚠️ 查询无效: {query}")
            return []
        resp.raise_for_status()
        return resp.json().get("items", [])
    except Exception as e:
        print(f"  ❌ 搜索失败: {e}")
        return []


def fetch_file_content(repo, path, ref=None):
    """获取文件原始内容（用于正则验证）"""
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    if ref:
        url += f"?ref={ref}"
    try:
        resp = requests.get(url, headers=github_headers(), timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("encoding") == "base64":
                import base64
                return base64.b64decode(data["content"]).decode("utf-8", errors="ignore")
    except:
        pass
    return ""


def is_placeholder(value):
    """判断是否为占位符"""
    lower = value.lower()
    return any(kw.lower() in lower for kw in PLACEHOLDER_KEYWORDS)


def extract_keys_from_text(text, pattern):
    """从文本中用正则提取密钥"""
    matches = re.findall(pattern, text)
    return [m for m in matches if not is_placeholder(m)]


def scan_all(max_results_per_pattern=20):
    """主扫描流程"""
    results = []
    seen = set()  # 去重: repo+path+provider

    total = len(SEARCH_PATTERNS)
    for idx, (query, (provider, key_regex, model)) in enumerate(SEARCH_PATTERNS, 1):
        print(f"[{idx}/{total}] 扫描 {provider} ... (query: {query})")

        items = search_github(query, per_page=min(max_results_per_pattern, 30))
        print(f"  → 命中 {len(items)} 个文件")

        for item in items:
            repo = item["repository"]["full_name"]
            file_path = item["path"]
            html_url = item.get("html_url", "")
            repo_url = item["repository"]["html_url"]
            owner = item["repository"]["owner"]["login"]
            is_private = item["repository"].get("private", False)
            ref = item.get("sha", "")

            dedup_key = f"{repo}|{file_path}|{provider}"
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            # 从 text_matches 中提取匹配片段
            text_matches = item.get("text_matches", [])
            fragment = ""
            for tm in text_matches:
                fragment += tm.get("fragment", "") + "\n"

            # 正则验证
            keys_found = extract_keys_from_text(fragment, key_regex)

            # 如果 fragment 里没匹配到，尝试拉文件内容
            if not keys_found and not is_private:
                content = fetch_file_content(repo, file_path, ref)
                if content:
                    keys_found = extract_keys_from_text(content, key_regex)

            if not keys_found:
                continue

            # 脱敏：只保留前 8 + 后 4
            for key in keys_found:
                masked = key[:8] + "..." + key[-4:] if len(key) > 16 else key[:4] + "..."
                results.append({
                    "provider": provider,
                    "model": model,
                    "repo": repo,
                    "owner": owner,
                    "file_path": file_path,
                    "repo_url": repo_url,
                    "file_url": html_url,
                    "key_preview": masked,
                    "is_private": is_private,
                    "scan_time": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })

        # GitHub 限速保护
        time.sleep(2)

    return results


def export_excel(results, output_path="github_keys_report.xlsx"):
    """导出 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "密钥扫描报告"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "序号", "供应商", "模型", "仓库", "所有者",
        "文件路径", "密钥预览", "仓库链接", "文件链接",
        "是否私有", "扫描时间"
    ]
    col_widths = [6, 18, 18, 30, 15, 40, 20, 40, 50, 10, 16]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    # 设置列宽
    from openpyxl.utils import get_column_letter
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 数据
    for row_idx, r in enumerate(results, 2):
        values = [
            row_idx - 1,
            r["provider"],
            r["model"],
            r["repo"],
            r["owner"],
            r["file_path"],
            r["key_preview"],
            r["repo_url"],
            r["file_url"],
            "是" if r["is_private"] else "否",
            r["scan_time"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # 高亮：私有仓库红色，公开仓库黄色
        fill = danger_fill if r["is_private"] else warn_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 统计 Sheet
    ws2 = wb.create_sheet("供应商统计")
    ws2.cell(row=1, column=1, value="供应商").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="暴露数量").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="涉及仓库数").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    from collections import Counter
    provider_count = Counter(r["provider"] for r in results)
    provider_repos = {}
    for r in results:
        provider_repos.setdefault(r["provider"], set()).add(r["repo"])

    for row, (provider, count) in enumerate(provider_count.most_common(), 2):
        ws2.cell(row=row, column=1, value=provider)
        ws2.cell(row=row, column=2, value=count)
        ws2.cell(row=row, column=3, value=len(provider_repos.get(provider, set())))

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    wb.save(output_path)
    return output_path


# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("🔍 GitHub AI 密钥扫描工具")
    print(f"⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    if not GITHUB_TOKEN:
        print("⚠️ 未设置 GITHUB_TOKEN，速率限制较低（10次/分钟）")
        print("   建议在脚本顶部填入 GitHub Token\n")

    results = scan_all(max_results_per_pattern=15)

    if results:
        out = export_excel(results)
        print(f"\n✅ 扫描完成！共发现 {len(results)} 个疑似密钥暴露")
        print(f"📊 报告已保存: {out}")

        # 打印摘要
        from collections import Counter
        print("\n📋 供应商分布:")
        for provider, count in Counter(r["provider"] for r in results).most_common():
            print(f"   {provider}: {count} 个")
    else:
        print("\n✅ 未发现密钥暴露（或 API 限速导致未完成）")
