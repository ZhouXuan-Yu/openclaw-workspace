"""从 scan_results.json 生成 Excel 报告"""
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 合并两轮结果
results = [
    {
        "provider": "DeepSeek",
        "model": "DeepSeek V2/V3",
        "repo": "Fenix19830717philip/zhongfei",
        "owner": "Fenix19830717philip",
        "file": "需求文档v1.txt",
        "key_preview": "sk-9e290...023f",
        "repo_url": "https://github.com/Fenix19830717philip/zhongfei",
        "file_url": "https://github.com/Fenix19830717philip/zhongfei/blob/main/需求文档v1.txt",
        "is_private": False,
        "scan_time": "2026-06-15 05:35",
    },
    {
        "provider": "DeepSeek",
        "model": "DeepSeek V2/V3",
        "repo": "chaocai1-lgtm/gzls_cc",
        "owner": "chaocai1-lgtm",
        "file": "config/history_config.py",
        "key_preview": "sk-bdf96...f5a9",
        "repo_url": "https://github.com/chaocai1-lgtm/gzls_cc",
        "file_url": "https://github.com/chaocai1-lgtm/gzls_cc/blob/main/config/history_config.py",
        "is_private": False,
        "scan_time": "2026-06-15 05:38",
    },
]

wb = Workbook()
ws = wb.active
ws.title = "密钥泄露报告"

# 样式
header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
normal_font = Font(name="微软雅黑", size=10)
link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 表头
headers = [
    "序号", "大模型供应商", "模型", "GitHub 仓库", "仓库所有者",
    "泄露文件路径", "密钥预览(脱敏)", "仓库链接", "文件链接(点击定位)",
    "是否私有仓库", "扫描时间"
]
col_widths = [6, 16, 16, 32, 18, 30, 18, 45, 55, 14, 16]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# 数据行
for row_idx, r in enumerate(results, 2):
    values = [
        row_idx - 1,
        r["provider"],
        r["model"],
        r["repo"],
        r["owner"],
        r["file"],
        r["key_preview"],
        r["repo_url"],
        r["file_url"],
        "是" if r["is_private"] else "否",
        r["scan_time"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # 链接列设为超链接
    repo_cell = ws.cell(row=row_idx, column=8)
    repo_cell.hyperlink = r["repo_url"]
    repo_cell.font = link_font

    file_cell = ws.cell(row=row_idx, column=9)
    file_cell.hyperlink = r["file_url"]
    file_cell.font = link_font

    # 高亮
    fill = danger_fill if r["is_private"] else warn_fill
    for col in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col).fill = fill

# 冻结首行 + 筛选
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# 行高
ws.row_dimensions[1].height = 25
for r in range(2, len(results) + 2):
    ws.row_dimensions[r].height = 22

# === Sheet 2: 说明 ===
ws2 = wb.create_sheet("扫描说明")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 60

info = [
    ("扫描时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("扫描关键词", "外贸、跨境、waimao、外贸客户、外贸开发信、外贸邮件"),
    ("覆盖供应商", "OpenAI / DeepSeek / Anthropic / Moonshot / DashScope / 智谱 / SiliconFlow / 通义千问 / 百川 / 讯飞 / MiniMax / 阶跃 / 零一 / 商汤 / 腾讯混元 / Groq / Replicate / Together / Fireworks / Perplexity / HuggingFace / Mistral / OpenRouter"),
    ("扫描方法", "GitHub Code Search API → 文件内容二次验证 → 正则提取"),
    ("密钥处理", "仅输出前8位+后4位脱敏预览，完整密钥见文件链接"),
    ("", ""),
    ("使用建议", ""),
    ("1. 通知仓库所有者", "在对应仓库开 Issue，告知 .env / 配置文件泄露了 API Key"),
    ("2. 脱敏沟通", "提供密钥预览即可，不需要完整密钥"),
    ("3. 建议操作", "所有者应立即轮换密钥 + 添加 .gitignore + 用 git filter-clean 清理历史"),
    ("", ""),
    ("列说明", ""),
    ("泄露文件路径", "泄露密钥的具体文件，所有者可直接定位修改"),
    ("文件链接", "GitHub 原始链接，点击可直接跳转到泄露代码位置"),
    ("密钥预览", "脱敏后的密钥片段，仅用于确认是自己的 key"),
]

for row, (key, val) in enumerate(info, 1):
    ws2.cell(row=row, column=1, value=key).font = Font(bold=True, name="微软雅黑", size=10)
    ws2.cell(row=row, column=2, value=val).font = normal_font

out_path = "github_keys_report.xlsx"
wb.save(out_path)
print(f"✅ Excel 已生成: {out_path}")
print(f"共 {len(results)} 条记录")
